"""
Configuration du projet DailyRH Scraper

Ce fichier centralise toutes les constantes et paramètres de configuration.
Modifier ce fichier permet d'adapter le comportement du scraper sans toucher au code.
"""

from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
from pathlib import Path
from datetime import date
# Racine du projet (un niveau au-dessus de /scripts)
BASE_DIR = Path(__file__).resolve().parent.parent




# ============================================================
# CONFIGURATION FICHIERS ET CHEMINS
# ============================================================

# Fichier de session SSO (contient les cookies d'authentification)
SESSION_FILE = BASE_DIR / "bnpparibas_session.json"

# Dossier output à la racine du projet
OUTPUT_DIR = BASE_DIR / "output"

# Noms des fichiers de sortie
OUTPUT_CSV = "extract_dailyRH.csv"
OUTPUT_EXCEL = "rapport_dailyRH.xlsx"

# ============================================================
# CONFIGURATION SCRAPING
# ============================================================

# URL de DailyRH
DAILYRH_URL = "https://dailyrh.hr.bnpparibas/app/foryou/#/demarches/teamplanning"

# Année à extraire
TARGET_YEAR = date.today().year

# Mode headless (True = pas d'interface graphique, False = navigateur visible)
HEADLESS_MODE = False

# Délais et timeouts (en secondes)
PAGE_LOAD_TIMEOUT = 10000  # Timeout de chargement de page (ms)
NAVIGATION_DELAY = 1.5     # Délai entre chaque changement de mois
INITIAL_LOAD_DELAY = 10    # Délai d'attente initial après ouverture de DailyRH
MAX_NAVIGATION_CLICKS = 50 # Nombre maximum de clics pour atteindre janvier

# ============================================================
# RÈGLES RH (PÉRIODE DE VÉRIFICATION)
# ============================================================

# Période de vérification des règles RH
RULE_START_DATE = datetime(2026, 5, 15)  # 15 mai 2026
RULE_END_DATE = datetime(2026, 10, 15)   # 15 octobre 2026

# Règles à vérifier
RULE_MIN_CONSECUTIVE_DAYS = 10  # Minimum 10 jours consécutifs de congés
RULE_MIN_TOTAL_DAYS = 20        # Minimum 20 jours de congés au total

# ============================================================
# CONSTANTES CALENDRIER
# ============================================================

# Abréviations des jours de la semaine (pour les en-têtes Excel)
JOURS_SEMAINE = ["L", "M", "Me", "J", "V", "S", "D"]

# Noms des mois en français
MOIS_NOMS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]

# Nombre de jours par mois en 2026 (pas bissextile)
#JOURS_PAR_MOIS_2026 = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

# Mapping des noms de mois français vers leur numéro
FRENCH_MONTHS = {
    "janvier": 1, "février": 2, "mars": 3, "avril": 4,
    "mai": 5, "juin": 6, "juillet": 7, "août": 8,
    "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12
}

# ============================================================
# STYLES EXCEL - COULEURS
# ============================================================

# En-têtes
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# Lignes de noms et totaux
NAME_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

# Indicateurs de conformité
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Mois dans les calendriers individuels
MONTH_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

# Types d'événements
TV_FILL = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")  # Télétravail validé
TP_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Télétravail à valider
CV_FILL = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")  # Congés validés
CP_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Congés à valider
RV_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # RTT validés
RP_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # RTT à valider
WE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Week-end/Férié
MIXED_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Journée mixte
INEX_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Jour inexistant

# ============================================================
# STYLES EXCEL - BORDURES
# ============================================================

# Bordure standard
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Bordure diagonale (pour les jours inexistants dans les calendriers)
BORDER_DIAG = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    diagonal=Side(style='thin', color='FF0000'),
    diagonalDown=True
)

# ============================================================
# STYLES EXCEL - DIMENSIONS
# ============================================================

# Largeurs de colonnes
EXCEL_COLUMN_WIDTHS = {
    'collaborateur': 25,     # Colonne nom dans feuille Synthèse
    'uid': 10,               # Colonne UID
    'metrics': 14,           # Colonnes de métriques (congés, RTT, etc.)
    'rules': 12,             # Colonnes de règles RH
    'month_name': 24,        # Colonne nom de mois dans feuilles mensuelles
    'day': 5.5,              # Colonnes de jours dans feuilles mensuelles
    'calendar_month': 12,    # Colonne mois dans calendriers individuels
    'calendar_day': 5        # Colonnes jours dans calendriers individuels
}

# Hauteurs de lignes
EXCEL_ROW_HEIGHTS = {
    'title': 25,    # Ligne de titre
    'legend': 22,   # Lignes de légende
    'header': 18,   # Lignes d'en-tête
    'data': None,   # Lignes de données (auto)
    'note': 30      # Lignes de notes
}
