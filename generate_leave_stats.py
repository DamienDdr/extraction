#!/usr/bin/env python3
"""
Version ULTRA-COMPATIBLE pour Excel 365 Entreprise
- Feuille Synthèse
- Une feuille par MOIS (collaborateurs en lignes, jours en colonnes, totaux en bas)
- Une feuille par COLLABORATEUR (calendrier annuel)
"""

import pandas as pd
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar


# ============================================================
# STYLES GLOBAUX
# ============================================================
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
NAME_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MONTH_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

TV_FILL = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
TP_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
CV_FILL = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
CP_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RV_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
RP_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
MIXED_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
INEX_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
BORDER_DIAG = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
    diagonal=Side(style='thin', color='FF0000'), diagonalDown=True
)

JOURS_SEMAINE = ["L", "M", "Me", "J", "V", "S", "D"]
MOIS_NOMS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]
JOURS_PAR_MOIS_2026 = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


# ============================================================
# FONCTIONS UTILITAIRES
# ============================================================

def is_validated(detail):
    if pd.isna(detail) or detail == '':
        return None
    detail_lower = str(detail).lower()
    if 'validé' in detail_lower or '(validé)' in detail_lower:
        return True
    elif 'à valider' in detail_lower or 'a valider' in detail_lower:
        return False
    return None


def is_rtt(detail):
    if pd.isna(detail) or detail == '':
        return False
    return 'rtt' in str(detail).lower()


def get_status_code(type_am, type_pm, detail_am, detail_pm):
    if type_am == 'JOUR_NON_OUVRE' and type_pm == 'JOUR_NON_OUVRE':
        return 'W'
    if type_am == 'PRESENT' and type_pm == 'PRESENT':
        return ''

    def half(type_val, detail_val):
        if type_val == 'PRESENT':
            return 'P'
        elif type_val == 'TELETRAVAIL':
            return 'TV' if is_validated(detail_val) else 'TP'
        elif type_val == 'CONGES':
            if is_rtt(detail_val):
                return 'RV' if is_validated(detail_val) else 'RP'
            else:
                return 'CV' if is_validated(detail_val) else 'CP'
        elif type_val == 'JOUR_NON_OUVRE':
            return 'W'
        return ''

    code_am = half(type_am, detail_am)
    code_pm = half(type_pm, detail_pm)

    if code_am == code_pm:
        return '' if code_am == 'P' else code_am

    if code_am == 'P' and code_pm != '':
        return f'{code_pm}-PM'
    if code_pm == 'P' and code_am != '':
        return f'{code_am}-AM'

    if code_am != '' and code_pm != '':
        return f'{code_am}/{code_pm}'

    return code_am if code_am else code_pm


def count_event_weight(code, prefixes=None):
    """
    Compte le poids d'un code en tenant compte des demi-journées.
    - "CV" → 1 (journée complète)
    - "CV-AM" ou "CV-PM" → 0.5 (demi-journée)
    - "CV/TV" → 0.5 pour CV, 0.5 pour TV
    - "W" ou "" → 0

    Si prefixes est None : compte tout événement (hors W et vide)
    Si prefixes est une liste : compte uniquement les codes qui matchent
    """
    if not code or code == '' or code == 'W':
        return 0

    # Cas journée mixte : "CV/TV", "RV/TP", etc.
    if '/' in code:
        parts = code.split('/')
        total = 0
        for part in parts:
            p = part.strip()
            if prefixes is None:
                total += 0.5
            else:
                for prefix in prefixes:
                    if p.startswith(prefix):
                        total += 0.5
                        break
        return total

    # Cas demi-journée : "CV-AM", "TV-PM", etc.
    if '-AM' in code or '-PM' in code:
        base = code.replace('-AM', '').replace('-PM', '')
        if prefixes is None:
            return 0.5
        for prefix in prefixes:
            if base.startswith(prefix):
                return 0.5
        return 0

    # Cas journée complète : "CV", "TV", "RV", etc.
    if prefixes is None:
        return 1
    for prefix in prefixes:
        if code.startswith(prefix):
            return 1
    return 0


def apply_cell_style(cell, code, is_even_row=False):
    """Applique le style visuel à une cellule selon le code"""
    cell.border = BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=7, bold=True)
    cell.value = code

    if code == 'W':
        cell.fill = WE_FILL
        cell.font = Font(size=7, color="999999")
    elif '/' in code and 'W' not in code:
        cell.fill = MIXED_FILL
    elif code.startswith('TV'):
        cell.fill = TV_FILL
    elif code.startswith('TP'):
        cell.fill = TP_FILL
    elif code.startswith('CV'):
        cell.fill = CV_FILL
    elif code.startswith('CP'):
        cell.fill = CP_FILL
    elif code.startswith('RV'):
        cell.fill = RV_FILL
    elif code.startswith('RP'):
        cell.fill = RP_FILL
    elif code == '' and is_even_row:
        cell.fill = NAME_FILL


def write_legend(ws, max_col_letter):
    """Écrit la légende sur les lignes 1-3"""
    ws.merge_cells(f'A1:{max_col_letter}1')
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = "LÉGENDE"
    ws['A2'].font = Font(bold=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = ""

    items_r2 = [
        ("B2:C2", "TV = Télétravail validé", TV_FILL, "FFFFFF"),
        ("D2:E2", "TP = Télétravail à valider", TP_FILL, "000000"),
        ("F2:G2", "CV = Congés validés", CV_FILL, "FFFFFF"),
        ("H2:I2", "CP = Congés à valider", CP_FILL, "000000"),
    ]
    items_r3 = [
        ("B3:C3", "RV = RTT validés", RV_FILL, "FFFFFF"),
        ("D3:E3", "RP = RTT à valider", RP_FILL, "000000"),
        ("F3:G3", "W = Week-end/Férié", WE_FILL, "000000"),
        ("H3:I3", "AM/PM = Demi-journée", MIXED_FILL, "000000"),
    ]

    for items in [items_r2, items_r3]:
        for cell_range, text, fill, fc in items:
            start = cell_range.split(':')[0]
            cell = ws[start]
            cell.value = text
            cell.fill = fill
            cell.font = Font(bold=True, size=8, color=fc)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
            ws.merge_cells(cell_range)

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22


# ============================================================
# ANALYSE DES DONNÉES
# ============================================================

def analyze_leave_data(csv_file):
    df = pd.read_csv(csv_file)
    df['date_obj'] = pd.to_datetime(df['date'], format='%Y/%m/%d')

    stats = defaultdict(lambda: {
        'teletravail_valide_am': 0, 'teletravail_valide_pm': 0,
        'teletravail_a_valider_am': 0, 'teletravail_a_valider_pm': 0,
        'conges_valides_am': 0, 'conges_valides_pm': 0,
        'conges_a_valider_am': 0, 'conges_a_valider_pm': 0,
        'rtt_valides_am': 0, 'rtt_valides_pm': 0,
        'rtt_a_valider_am': 0, 'rtt_a_valider_pm': 0,
        'regle_10j_consecutifs': False, 'regle_20j_total': False,
        'jours_consecutifs_max': 0, 'jours_total_periode': 0,
        'uid': '',
    })

    date_debut = datetime(2026, 5, 15)
    date_fin = datetime(2026, 10, 15)

    for _, row in df.iterrows():
        c = row['collaborateur']
        if 'uid' in row and row['uid'] and not stats[c]['uid']:
            stats[c]['uid'] = str(row['uid'])

        for period in ['am', 'pm']:
            t = row[f'type_{period}']
            d = row[f'detail_{period}']
            v = is_validated(d)
            if t == 'TELETRAVAIL':
                if v == True: stats[c][f'teletravail_valide_{period}'] += 1
                elif v == False: stats[c][f'teletravail_a_valider_{period}'] += 1
            elif t == 'CONGES':
                if is_rtt(d):
                    if v == True: stats[c][f'rtt_valides_{period}'] += 1
                    elif v == False: stats[c][f'rtt_a_valider_{period}'] += 1
                else:
                    if v == True: stats[c][f'conges_valides_{period}'] += 1
                    elif v == False: stats[c][f'conges_a_valider_{period}'] += 1

    for collaborateur in stats.keys():
        collab_df = df[df['collaborateur'] == collaborateur].copy()
        collab_df = collab_df[(collab_df['date_obj'] >= date_debut) & (collab_df['date_obj'] <= date_fin)]
        collab_df = collab_df.sort_values('date_obj')

        jours_type = {}
        total_jours = 0
        for _, row in collab_df.iterrows():
            dt = row['date_obj']
            ca = row['type_am'] == 'CONGES'
            cp = row['type_pm'] == 'CONGES'
            we = row['type_am'] == 'JOUR_NON_OUVRE' and row['type_pm'] == 'JOUR_NON_OUVRE'
            if ca or cp:
                jours_type[dt] = 'CONGES'
                total_jours += 1 if (ca and cp) else 0.5
            elif we:
                jours_type[dt] = 'WEEKEND'
            else:
                jours_type[dt] = 'AUTRE'

        max_cons = 0
        cur_cons = 0
        cur = date_debut
        while cur <= date_fin:
            jt = jours_type.get(cur, 'AUTRE')
            if jt == 'CONGES':
                cur_cons += 1
                max_cons = max(max_cons, cur_cons)
            elif jt != 'WEEKEND':
                cur_cons = 0
            cur += pd.Timedelta(days=1)

        stats[collaborateur]['jours_consecutifs_max'] = max_cons
        stats[collaborateur]['jours_total_periode'] = total_jours
        stats[collaborateur]['regle_10j_consecutifs'] = max_cons >= 10
        stats[collaborateur]['regle_20j_total'] = total_jours >= 20

    return stats


# ============================================================
# FEUILLE SYNTHÈSE
# ============================================================

def create_summary_sheet(wb, stats):
    ws = wb.active
    ws.title = "Synthèse"

    headers = [
        "Collaborateur", "UID",
        "Télétravail\nValidé (j)", "Télétravail\nÀ valider (j)",
        "Congés\nValidés (j)", "Congés\nÀ valider (j)",
        "RTT\nValidés (j)", "RTT\nÀ valider (j)",
        "Règle 10j\nconsécutifs", "Règle 20j\ntotal"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    sorted_collabs = sorted(stats.keys())
    row = 2
    for collaborateur in sorted_collabs:
        s = stats[collaborateur]
        data = [
            collaborateur, s['uid'],
            (s['teletravail_valide_am'] + s['teletravail_valide_pm']) / 2,
            (s['teletravail_a_valider_am'] + s['teletravail_a_valider_pm']) / 2,
            (s['conges_valides_am'] + s['conges_valides_pm']) / 2,
            (s['conges_a_valider_am'] + s['conges_a_valider_pm']) / 2,
            (s['rtt_valides_am'] + s['rtt_valides_pm']) / 2,
            (s['rtt_a_valider_am'] + s['rtt_a_valider_pm']) / 2,
            "✓" if s['regle_10j_consecutifs'] else "✗",
            "✓" if s['regle_20j_total'] else "✗"
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if col == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif 3 <= col <= 8:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(value, (int, float)):
                    cell.number_format = '0.0'
            elif col > 8:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True, size=14)
                cell.fill = GREEN_FILL if value == "✓" else RED_FILL
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2, value="").fill = SUBHEADER_FILL
    ws.cell(row=row, column=2).border = BORDER

    for col in range(3, 9):
        cl = get_column_letter(col)
        cell = ws.cell(row=row, column=col, value=f"=SUM({cl}2:{cl}{row - 1})")
        cell.font = Font(bold=True)
        cell.fill = SUBHEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '0.0'

    nb_ok_10j = sum(1 for s in stats.values() if s['regle_10j_consecutifs'])
    nb_ok_20j = sum(1 for s in stats.values() if s['regle_20j_total'])
    for col, val in [(9, f"{nb_ok_10j}/{len(stats)}"), (10, f"{nb_ok_20j}/{len(stats)}")]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = SUBHEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12

    ws.merge_cells(f'A{row + 2}:J{row + 2}')
    note = ws[f'A{row + 2}']
    note.value = "📋 Règles RH (période 15/05 - 15/10) : 10j consécutifs = au moins 10 jours d'affilée | 20j total = au moins 20 jours (consécutifs ou non)"
    note.font = Font(size=9, italic=True)
    note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row + 2].height = 30
    ws.freeze_panes = 'A2'


# ============================================================
# FEUILLES MENSUELLES
# ============================================================

def create_monthly_sheets(wb, csv_file):
    df = pd.read_csv(csv_file)
    df['date_obj'] = pd.to_datetime(df['date'], format='%Y/%m/%d')
    collaborateurs = sorted(df['collaborateur'].unique())
    months_in_data = sorted(df['date_obj'].dt.month.unique())

    for month_num in months_in_data:
        month_name = MOIS_NOMS[month_num - 1]
        nb_days = calendar.monthrange(2026, month_num)[1]

        ws = wb.create_sheet(month_name)
        print(f"  📅 Feuille {month_name}...")

        # Titre + légende
        ws['A1'].value = f"PLANNING {month_name.upper()} 2026"
        write_legend(ws, get_column_letter(nb_days + 1))

        # En-têtes ligne 5 (jour semaine) et ligne 6 (numéro)
        ws.cell(row=5, column=1, value="").fill = HEADER_FILL
        ws.cell(row=5, column=1).border = BORDER
        ws.cell(row=6, column=1, value="Collaborateur").fill = HEADER_FILL
        ws.cell(row=6, column=1).font = HEADER_FONT
        ws.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=6, column=1).border = BORDER

        for day in range(1, nb_days + 1):
            col = day + 1
            dow = calendar.weekday(2026, month_num, day)

            cell_dow = ws.cell(row=5, column=col, value=JOURS_SEMAINE[dow])
            cell_dow.alignment = Alignment(horizontal='center', vertical='center')
            cell_dow.font = Font(size=8, bold=True, color="FFFFFF")
            cell_dow.fill = HEADER_FILL
            cell_dow.border = BORDER

            cell_day = ws.cell(row=6, column=col, value=day)
            cell_day.alignment = Alignment(horizontal='center', vertical='center')
            cell_day.font = HEADER_FONT
            cell_day.fill = HEADER_FILL
            cell_day.border = BORDER

        ws.row_dimensions[5].height = 15
        ws.row_dimensions[6].height = 18

        # Données
        month_df = df[df['date_obj'].dt.month == month_num]
        data_start_row = 7

        for idx, collaborateur in enumerate(collaborateurs):
            row_num = data_start_row + idx
            collab_month = month_df[month_df['collaborateur'] == collaborateur]

            cell = ws.cell(row=row_num, column=1, value=collaborateur)
            cell.fill = NAME_FILL if idx % 2 == 0 else PatternFill()
            cell.font = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = BORDER

            day_codes = {}
            for _, r in collab_month.iterrows():
                d = r['date_obj'].day
                day_codes[d] = get_status_code(r['type_am'], r['type_pm'], r['detail_am'], r['detail_pm'])

            for day in range(1, nb_days + 1):
                col = day + 1
                cell = ws.cell(row=row_num, column=col)
                code = day_codes.get(day, '')
                apply_cell_style(cell, code, is_even_row=(idx % 2 == 0))

        # ============================================================
        # TOTAUX en bas (avec comptage correct des demi-journées)
        # ============================================================
        separator_row = data_start_row + len(collaborateurs)
        total_row = separator_row + 1

        for col in range(1, nb_days + 2):
            ws.cell(row=separator_row, column=col).border = Border(bottom=Side(style='medium'))

        total_labels = [
            ("TOTAL événements (j)", None),
            ("  dont Télétravail (j)", ["TV", "TP"]),
            ("  dont Congés (j)", ["CV", "CP"]),
            ("  dont RTT (j)", ["RV", "RP"]),
        ]

        for offset, (label, prefixes) in enumerate(total_labels):
            r = total_row + offset

            cell = ws.cell(row=r, column=1, value=label)
            cell.font = Font(bold=(offset == 0), size=9, italic=(offset > 0))
            cell.fill = TOTAL_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center')

            for day in range(1, nb_days + 1):
                col = day + 1
                total = 0

                for c_idx in range(len(collaborateurs)):
                    code_val = ws.cell(row=data_start_row + c_idx, column=col).value or ''
                    total += count_event_weight(code_val, prefixes)

                cell = ws.cell(row=r, column=col)
                if total > 0:
                    cell.value = int(total) if total == int(total) else total
                    cell.number_format = '0.0' if total != int(total) else '0'
                else:
                    cell.value = ""
                cell.font = Font(bold=(offset == 0), size=9)
                cell.fill = TOTAL_FILL
                cell.border = BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Mise en forme
        ws.column_dimensions['A'].width = 24
        for col in range(2, nb_days + 2):
            ws.column_dimensions[get_column_letter(col)].width = 5.5
        ws.freeze_panes = 'B7'


# ============================================================
# FEUILLES PAR COLLABORATEUR
# ============================================================

def create_calendar_sheets(wb, csv_file):
    df = pd.read_csv(csv_file)
    df['date_obj'] = pd.to_datetime(df['date'], format='%Y/%m/%d')
    collaborateurs = sorted(df['collaborateur'].unique())

    for collaborateur in collaborateurs:
        ws = wb.create_sheet(collaborateur[:31])
        print(f"  👤 Feuille {collaborateur}...")

        # Titre + légende
        ws['A1'].value = f"PLANNING 2026 - {collaborateur}"
        write_legend(ws, 'AF')

        # Explication demi-journées
        ws.merge_cells('A4:I4')
        ws['A4'].value = "💡 Code-AM/PM = demi-journée | Code1/Code2 = matin≠après-midi"
        ws['A4'].font = Font(size=9, italic=True)
        ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A4'].fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        ws.row_dimensions[4].height = 20

        # En-têtes
        ws.cell(row=6, column=1, value="Mois").fill = HEADER_FILL
        ws.cell(row=6, column=1).font = HEADER_FONT
        ws.cell(row=6, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=6, column=1).border = BORDER

        for day in range(1, 32):
            col = day + 1
            cell = ws.cell(row=6, column=col, value=day)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        ws.row_dimensions[6].height = 20

        ws.column_dimensions['A'].width = 12
        for col in range(2, 33):
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Données
        collab_df = df[df['collaborateur'] == collaborateur].copy()
        collab_data = {}
        for _, row in collab_df.iterrows():
            m = row['date_obj'].month
            d = row['date_obj'].day
            if m not in collab_data:
                collab_data[m] = {}
            collab_data[m][d] = get_status_code(row['type_am'], row['type_pm'], row['detail_am'], row['detail_pm'])

        for month_num, month_name in enumerate(MOIS_NOMS, 1):
            row_num = 7 + month_num - 1
            max_days = JOURS_PAR_MOIS_2026[month_num - 1]

            cell = ws.cell(row=row_num, column=1, value=month_name)
            cell.fill = MONTH_FILL
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER

            for day in range(1, 32):
                col = day + 1
                cell = ws.cell(row=row_num, column=col)

                if day <= max_days:
                    code = collab_data.get(month_num, {}).get(day, '')
                    apply_cell_style(cell, code)
                else:
                    cell.border = BORDER_DIAG
                    cell.fill = INEX_FILL

        ws.freeze_panes = 'B7'


# ============================================================
# ASSEMBLAGE FINAL
# ============================================================

def create_excel_report(stats, csv_file, output_file):
    wb = openpyxl.Workbook()

    print("📊 Création de la feuille Synthèse...")
    create_summary_sheet(wb, stats)

    print("📅 Création des feuilles mensuelles...")
    create_monthly_sheets(wb, csv_file)

    print("👤 Création des feuilles par collaborateur...")
    create_calendar_sheets(wb, csv_file)

    wb.save(output_file)
    print(f"✅ Fichier Excel créé : {output_file}")


def main():
    input_file = "leave_planning_2026_complete.csv"
    output_file = "compteurs_conges_2026_compatible.xlsx"

    print("🔍 Analyse du fichier CSV...")
    stats = analyze_leave_data(input_file)
    print(f"📊 Nombre de collaborateurs : {len(stats)}")

    print("📝 Génération du fichier Excel...")
    create_excel_report(stats, input_file, output_file)

    print("\n✨ Terminé !")
    print("\nContenu du fichier :")
    print("  - Feuille 'Synthèse' : compteurs + règles RH")
    print("  - Une feuille par mois (collaborateurs en lignes, totaux en bas)")
    print("  - Une feuille par collaborateur (calendrier annuel)")


if __name__ == "__main__":
    main()