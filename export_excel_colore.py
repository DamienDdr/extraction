import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT = "leave_planning.csv"
OUTPUT = "leave_planning_colore.xlsx"

# -----------------------
# Chargement CSV
# -----------------------
df = pd.read_csv(INPUT)

def extract_date(classes):
    if not isinstance(classes, str):
        return None
    m = re.search(r"(20\d{2}/\d{2}/\d{2})", classes)
    return m.group(1) if m else None

def classify(classes, title):
    classes = "" if pd.isna(classes) else str(classes)
    title = "" if pd.isna(title) else str(title)

    if "grey_cell_weekend" in classes:
        return "WEEKEND"
    if "telework" in classes:
        return "TELETRAVAIL"
    if "Congés" in title:
        return "CONGES"
    if "validated_vcell" in classes:
        return "PRESENT"
    return "INCONNU"

df["date"] = df["classes"].apply(extract_date)
df["type"] = df.apply(lambda r: classify(r["classes"], r["title"]), axis=1)

df = df[["collaborateur", "date", "type", "title"]]
df.sort_values(["collaborateur", "date"], inplace=True)

# -----------------------
# Export Excel
# -----------------------
df.to_excel(OUTPUT, index=False)

wb = load_workbook(OUTPUT)
ws = wb.active

# -----------------------
# Styles
# -----------------------
FILLS = {
    "TELETRAVAIL": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "CONGES": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "PRESENT": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    "WEEKEND": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
}

for row in range(2, ws.max_row + 1):
    cell_type = ws[f"C{row}"].value
    fill = FILLS.get(cell_type)
    if fill:
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].fill = fill

# Ajustement largeur colonnes
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 35

wb.save(OUTPUT)

print(f"✅ Export Excel coloré généré → {OUTPUT}")
